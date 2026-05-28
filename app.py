from flask import Flask, request, render_template_string, send_file
import os
from werkzeug.utils import secure_filename
from openpyxl import Workbook, load_workbook
from reportlab.pdfgen import canvas
from twilio.rest import Client
from datetime import datetime

app = Flask(__name__)

# =========================
# FOLDERS
# =========================

UPLOAD_FOLDER = "static/uploads"
PDF_FOLDER = "receipts"

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(PDF_FOLDER, exist_ok=True)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# =========================
# EXCEL FILE
# =========================

EXCEL_FILE = "students.xlsx"

if not os.path.exists(EXCEL_FILE):

    wb = Workbook()
    ws = wb.active

    ws.append([
        "Student Name",
        "Father Name",
        "Mother Name",
        "Mobile",
        "Aadhaar",
        "Address",
        "Course",
        "Qualification",
        "Photo"
    ])

    wb.save(EXCEL_FILE)

# =========================
# TWILIO WHATSAPP
# =========================

TWILIO_SID = os.environ.get("TWILIO_SID")
TWILIO_TOKEN = os.environ.get("TWILIO_TOKEN")
TWILIO_WHATSAPP = os.environ.get("TWILIO_WHATSAPP")
YOUR_WHATSAPP = os.environ.get("YOUR_WHATSAPP")

client = None

if TWILIO_SID and TWILIO_TOKEN:
    client = Client(TWILIO_SID, TWILIO_TOKEN)

# =========================
# ALLOWED IMAGE TYPES
# =========================

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg'}

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# =========================
# HTML
# =========================

HTML = """

<!DOCTYPE html>
<html>

<head>

<title>CT Institute Admission Form</title>

<meta name="viewport" content="width=device-width, initial-scale=1.0">

<style>

body{
    margin:0;
    padding:0;
    font-family:Arial,sans-serif;
    background:linear-gradient(135deg,#ff9a9e,#fad0c4,#fbc2eb,#a18cd1);
    min-height:100vh;
}

.container{
    width:90%;
    max-width:650px;
    background:white;
    margin:25px auto;
    padding:30px;
    border-radius:20px;
    box-shadow:0 0 20px rgba(0,0,0,0.3);
}

.logo{
    width:160px;
    height:160px;
    border-radius:50%;
    background:black;
    border:6px solid gold;
    margin:auto;
    display:flex;
    align-items:center;
    justify-content:center;
    font-size:70px;
    font-weight:bold;
    color:gold;
}

h1{
    text-align:center;
    margin-top:15px;
    color:#111;
}

.title{
    text-align:center;
    color:#555;
    margin-bottom:25px;
    line-height:28px;
}

label{
    font-weight:bold;
    display:block;
    margin-top:15px;
    margin-bottom:5px;
}

input,
textarea,
select{
    width:100%;
    padding:14px;
    border-radius:10px;
    border:2px solid #ddd;
    box-sizing:border-box;
    font-size:16px;
}

textarea{
    height:100px;
    resize:none;
}

.submit-btn{
    width:100%;
    background:linear-gradient(90deg,#ff512f,#dd2476);
    color:white;
    border:none;
    padding:15px;
    border-radius:12px;
    font-size:20px;
    margin-top:25px;
    cursor:pointer;
    font-weight:bold;
}

.submit-btn:hover{
    opacity:0.9;
}

.footer{
    text-align:center;
    margin-top:20px;
    color:#444;
}

</style>

</head>

<body>

<div class="container">

<div class="logo">CT</div>

<h1>CT Institute</h1>

<div class="title">
Online Admission Form <br>
Admission Through Vikash Maurya <br>
Location: Deoria, Uttar Pradesh
</div>

<form method="POST" enctype="multipart/form-data">

<label>Student Name</label>
<input type="text" name="student_name" required>

<label>Father Name</label>
<input type="text" name="father_name" required>

<label>Mother Name</label>
<input type="text" name="mother_name" required>

<label>Mobile Number</label>
<input type="text" name="mobile" required>

<label>Aadhaar Number</label>
<input type="text" name="aadhaar" required>

<label>Address</label>
<textarea name="address" required></textarea>

<label>Select Course</label>

<select name="course">

<option>ADCA</option>
<option>DCA</option>
<option>ADFA</option>
<option>CCC</option>
<option>C#</option>
<option>C++</option>
<option>O LEVEL</option>
<option>TALLY</option>
<option>DTP</option>
<option>EXCEL</option>
<option>ALL COMPUTER COURSE</option>

</select>

<label>Qualification</label>
<input type="text" name="qualification" required>

<label>Upload Student Photo</label>
<input type="file" name="photo" required>

<button class="submit-btn" type="submit">
Submit Admission Form
</button>

</form>

<div class="footer">
© CT Institute | Powered By Vikash Maurya
</div>

</div>

</body>
</html>

"""

# =========================
# HOME ROUTE
# =========================

@app.route('/', methods=['GET', 'POST'])

def home():

    if request.method == 'POST':

        try:

            student_name = request.form['student_name']
            father_name = request.form['father_name']
            mother_name = request.form['mother_name']
            mobile = request.form['mobile']
            aadhaar = request.form['aadhaar']
            address = request.form['address']
            course = request.form['course']
            qualification = request.form['qualification']

            # =========================
            # PHOTO SAVE
            # =========================

            photo = request.files['photo']

            if photo.filename == '':
                return "No file selected"

            if not allowed_file(photo.filename):
                return "Only PNG, JPG, JPEG allowed"

            filename = secure_filename(photo.filename)

            unique_name = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{filename}"

            photo_path = os.path.join(
                app.config['UPLOAD_FOLDER'],
                unique_name
            )

            photo.save(photo_path)

            # =========================
            # SAVE TO EXCEL
            # =========================

            wb = load_workbook(EXCEL_FILE)
            ws = wb.active

            ws.append([
                student_name,
                father_name,
                mother_name,
                mobile,
                aadhaar,
                address,
                course,
                qualification,
                unique_name
            ])

            wb.save(EXCEL_FILE)

            # =========================
            # PDF RECEIPT
            # =========================

            pdf_name = f"{student_name}_{mobile}.pdf"

            pdf_path = os.path.join(PDF_FOLDER, pdf_name)

            c = canvas.Canvas(pdf_path)

            c.setFont("Helvetica-Bold", 20)
            c.drawString(160, 800, "CT Institute Receipt")

            c.setFont("Helvetica", 14)

            c.drawString(80, 740, f"Student Name: {student_name}")
            c.drawString(80, 710, f"Father Name: {father_name}")
            c.drawString(80, 680, f"Mother Name: {mother_name}")
            c.drawString(80, 650, f"Mobile: {mobile}")
            c.drawString(80, 620, f"Course: {course}")
            c.drawString(80, 590, f"Qualification: {qualification}")

            c.save()

            # =========================
            # WHATSAPP MESSAGE
            # =========================

            if client:

                try:

                    msg = f'''
New Admission Received

Name: {student_name}
Mobile: {mobile}
Course: {course}
'''

                    client.messages.create(
                        from_=TWILIO_WHATSAPP,
                        body=msg,
                        to=YOUR_WHATSAPP
                    )

                except Exception as e:
                    print("WhatsApp Error:", e)

            return send_file(pdf_path, as_attachment=True)

        except Exception as e:
            return f"Error: {str(e)}"

    return render_template_string(HTML)

# =========================
# ADMIN PANEL
# =========================

@app.route('/admin')

def admin():

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    data = list(ws.values)

    html = """

    <html>

    <head>

    <title>Admin Panel</title>

    <style>

    body{
        font-family:Arial;
        background:#f2f2f2;
        padding:20px;
    }

    table{
        width:100%;
        border-collapse:collapse;
        background:white;
    }

    th,td{
        border:1px solid gray;
        padding:10px;
        text-align:center;
    }

    th{
        background:black;
        color:white;
    }

    img{
        width:70px;
        height:70px;
        border-radius:10px;
    }

    </style>

    </head>

    <body>

    <h1>CT Institute Admin Panel</h1>

    <table>

    """

    for i, row in enumerate(data):

        html += "<tr>"

        for j, col in enumerate(row):

            if i != 0 and j == 8:
                html += f"<td><img src='/static/uploads/{col}'></td>"
            else:
                tag = "th" if i == 0 else "td"
                html += f"<{tag}>{col}</{tag}>"

        html += "</tr>"

    html += """

    </table>

    </body>
    </html>

    """

    return html

# =========================
# RUN
# =========================

if __name__ == "__main__":

    port = int(os.environ.get("PORT", 10000))

    app.run(
        host="0.0.0.0",
        port=port
    )